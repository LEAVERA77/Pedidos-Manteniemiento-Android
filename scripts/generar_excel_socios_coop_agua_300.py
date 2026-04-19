# -*- coding: utf-8 -*-
"""
Genera Excel de 300 socios ficticios para cooperativa de agua (Entre Ríos).
Usa Nominatim (1 petición ~1.1 s): centro por localidad + reverse geocode
para direcciones existentes en OSM.

Columnas: nis, medidor, nombre, calle, numero, localidad, latitud, longitud.

made by leavera77
"""
from __future__ import annotations

import random
import time
from typing import Any, Dict, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

NOM_SEARCH = "https://nominatim.openstreetmap.org/search"
NOM_REVERSE = "https://nominatim.openstreetmap.org/reverse"
USER_AGENT = "GestorNova-cooperativa-agua/1.0 (generacion-interna-socios)"

LOCALIDADES_FILAS = [
    ("Hasenkamp", 75),
    ("María Grande", 75),
    ("Cerrito", 75),
    ("El Pingo", 75),
]

NOMBRES = [
    "Laura", "Mariana", "Silvia", "Gabriela", "Valeria", "Andrea", "Paula",
    "Natalia", "Carolina", "Verónica", "Patricia", "Lucía", "Sofía", "María",
    "Alejandro", "Martín", "Diego", "Federico", "Roberto", "Daniel",
    "Gustavo", "Miguel", "Rodrigo", "Lucas", "Facundo", "Fabricio", "Javier",
]
APELLIDOS = [
    "Acosta", "Benítez", "Corradi", "Domínguez", "Estévez", "Figueroa",
    "González", "Herrera", "Ibáñez", "Juárez", "Luna", "Molina",
    "Nuñez", "Ocampo", "Peretti", "Quinteros", "Ramírez", "Sosa", "Torres",
    "Vega", "Zárate", "Aguirre", "Basualdo", "Cabral",
]


def espera() -> None:
    time.sleep(1.12)


def nominatim_centro(localidad: str) -> Tuple[float, float]:
    params = {
        "q": f"{localidad}, Entre Ríos, Argentina",
        "format": "json",
        "limit": 1,
        "countrycodes": "ar",
    }
    r = requests.get(
        NOM_SEARCH,
        params=params,
        headers={"User-Agent": USER_AGENT, "Accept-Language": "es"},
        timeout=40,
    )
    r.raise_for_status()
    data = r.json()
    espera()
    if not data:
        raise RuntimeError(f"No se encontró centro para: {localidad}")
    return float(data[0]["lat"]), float(data[0]["lon"])


def nominatim_reverse(lat: float, lon: float) -> Dict[str, Any]:
    r = requests.get(
        NOM_REVERSE,
        params={
            "lat": lat,
            "lon": lon,
            "format": "json",
            "addressdetails": 1,
            "zoom": 18,
        },
        headers={"User-Agent": USER_AGENT, "Accept-Language": "es"},
        timeout=40,
    )
    r.raise_for_status()
    espera()
    return r.json()


def extraer_calle_numero(addr: Dict[str, Any]) -> Tuple[str, str]:
    """Extrae calle y número desde bloque address de Nominatim."""
    a = addr.get("address") or {}
    road = (
        a.get("road")
        or a.get("pedestrian")
        or a.get("path")
        or a.get("footway")
        or a.get("residential")
        or ""
    )
    house = (
        a.get("house_number")
        or a.get("house_name")
        or ""
    )
    if isinstance(road, str):
        road = road.strip()
    else:
        road = ""
    if isinstance(house, str):
        house = house.strip()
    else:
        house = str(house) if house else ""

    if not road:
        road = "Sin nombre de vía OSM"
    if not house:
        house = "s/n"
    return road[:80], house[:20]


def nombre_ficticio(i: int) -> str:
    rng = random.Random(i * 7919 + 424242)
    n = rng.choice(NOMBRES)
    a = rng.choice(APELLIDOS)
    b = rng.choice(APELLIDOS)
    while b == a:
        b = rng.choice(APELLIDOS)
    return f"{n} {a} {b}"


def punto_aleatorio_cerca(lat0: float, lon0: float, radio_grados: float) -> Tuple[float, float]:
    """Pequeño desplazamiento aleatorio (~ hasta unos km según latitud)."""
    return (
        lat0 + random.uniform(-radio_grados, radio_grados),
        lon0 + random.uniform(-radio_grados, radio_grados),
    )


def main() -> None:
    random.seed(20260416)
    centros: Dict[str, Tuple[float, float]] = {}
    for loc, _ in LOCALIDADES_FILAS:
        centros[loc] = nominatim_centro(loc)

    filas: list[dict] = []
    nis_base = 37250001

    # Radio ~0.025° lat ≈ 2–3 km según zona (suficiente para caer en vías urbanas/rurales)
    radio = 0.028

    idx = 0
    for localidad, cantidad in LOCALIDADES_FILAS:
        lat0, lon0 = centros[localidad]
        intentos = 0
        generadas = 0
        while generadas < cantidad:
            intentos += 1
            if intentos > cantidad * 25:
                raise RuntimeError(f"Demasiados reintentos en {localidad}")
            lat, lon = punto_aleatorio_cerca(lat0, lon0, radio)
            try:
                rev = nominatim_reverse(lat, lon)
            except Exception:
                continue
            if rev.get("error"):
                continue
            calle, numero = extraer_calle_numero(rev)
            # Aceptar cualquier resultado de reverse (campo rural puede no tener «road» en OSM).

            idx += 1
            generadas += 1
            filas.append(
                {
                    "nis": nis_base + idx - 1,
                    "medidor": str(915000000 + idx),
                    "nombre": nombre_ficticio(idx),
                    "calle": calle,
                    "numero": numero,
                    "localidad": localidad,
                    "latitud": round(lat, 7),
                    "longitud": round(lon, 7),
                }
            )

    assert len(filas) == 300

    wb = Workbook()
    ws = wb.active
    ws.title = "socios"
    headers = [
        "nis",
        "medidor",
        "nombre",
        "calle",
        "numero",
        "localidad",
        "latitud",
        "longitud",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(filas, start=2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    out_path = (
        "app/src/main/assets/ejemplos/socios-cooperativa-agua-300-entre-rios.xlsx"
    )
    wb.save(out_path)
    print(f"OK: {out_path} ({len(filas)} filas)")


if __name__ == "__main__":
    main()
