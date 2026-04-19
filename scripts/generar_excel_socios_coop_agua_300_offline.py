# -*- coding: utf-8 -*-
"""
Genera Excel 300 socios cooperativa de agua (Entre Ríos), 75 por localidad.
Coordenadas: puntos aleatorios WGS84 dentro de cajas ~municipio (aprox. OSM);
calle/número: combinación ficticia plausible (sin «demo» en nombres).
Importación GestorNova: nis, medidor, nombre, calle, numero, localidad, latitud, longitud.

made by leavera77
"""
from __future__ import annotations

import random
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Cajas de referencia (aprox. extensión urbana-rural ~ Entre Ríos; ajustables)
# Hasenkamp, María Grande, Cerrito (zona Paraná), El Pingo
CAJAS: Dict[str, Tuple[float, float, float, float]] = {
    "Hasenkamp": (-31.535, -31.490, -58.655, -58.595),
    "María Grande": (-31.695, -31.645, -58.730, -58.680),
    "Cerrito": (-31.600, -31.550, -58.090, -58.020),
    "El Pingo": (-31.445, -31.385, -58.815, -58.750),
}

CALLES: List[str] = [
    "San Martín",
    "Belgrano",
    "Mitre",
    "Sarmiento",
    "Rivadavia",
    "Urquiza",
    "Alvear",
    "French",
    "Rosas",
    "9 de Julio",
    "25 de Mayo",
    "Las Heras",
    "Independencia",
    "Paraná",
    "Buenos Aires",
    "Santa Fe",
    "Lavalle",
    "Güemes",
    "Castelli",
    "Illia",
]

NOMBRES = [
    "Laura", "Mariana", "Silvia", "Gabriela", "Valeria", "Andrea", "Paula",
    "Natalia", "Carolina", "Lucía", "Sofía", "María", "Romina", "Daniela",
    "Alejandro", "Martín", "Diego", "Federico", "Roberto", "Daniel",
    "Gustavo", "Miguel", "Rodrigo", "Lucas", "Facundo", "Javier",
]
APELLIDOS = [
    "Acosta", "Benítez", "Corradi", "Domínguez", "Estévez", "Figueroa",
    "González", "Herrera", "Ibáñez", "Juárez", "Luna", "Molina", "Nuñez",
    "Ocampo", "Peretti", "Quinteros", "Ramírez", "Sosa", "Torres", "Vega",
    "Aguirre", "Basualdo", "Cabral", "Dure", "Etcheverry",
]

LOCALIDADES_ORDEN = [
    ("Hasenkamp", 75),
    ("María Grande", 75),
    ("Cerrito", 75),
    ("El Pingo", 75),
]


def nombre_ficticio(i: int) -> str:
    r = random.Random(i * 11003 + 17)
    n = r.choice(NOMBRES)
    a = r.choice(APELLIDOS)
    b = r.choice(APELLIDOS)
    while b == a:
        b = r.choice(APELLIDOS)
    return f"{n} {a} {b}"


def random_en_caja(lat_s: float, lat_n: float, lon_w: float, lon_e: float) -> Tuple[float, float]:
    lat = random.uniform(lat_s, lat_n)
    lon = random.uniform(lon_w, lon_e)
    return round(lat, 7), round(lon, 7)


def main() -> None:
    random.seed(20260416)
    filas = []
    nis0 = 37290001
    idx = 0

    for localidad, cantidad in LOCALIDADES_ORDEN:
        lat_s, lat_n, lon_w, lon_e = CAJAS[localidad]
        for _ in range(cantidad):
            idx += 1
            lat, lon = random_en_caja(lat_s, lat_n, lon_w, lon_e)
            calle = random.choice(CALLES)
            numero = str(random.randint(12, 2899))
            filas.append(
                {
                    "nis": nis0 + idx - 1,
                    "medidor": str(916000000 + idx),
                    "nombre": nombre_ficticio(idx),
                    "calle": calle,
                    "numero": numero,
                    "localidad": localidad,
                    "latitud": lat,
                    "longitud": lon,
                }
            )

    assert len(filas) == 300

    wb = Workbook()
    ws = wb.active
    ws.title = "socios"
    headers = [
        "nis",
        "medidor",
        "nombre",
        "calle",
        "numero",
        "localidad",
        "latitud",
        "longitud",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(filas, start=2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    out = "app/src/main/assets/ejemplos/socios-cooperativa-agua-300-entre-rios.xlsx"
    wb.save(out)
    print(f"OK {out} filas={len(filas)}")


if __name__ == "__main__":
    main()
